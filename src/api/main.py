"""FastAPI application — PM Agent + Chaos Testing Agent combined API."""

import os
from contextlib import asynccontextmanager
from typing import Any

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from src.agent.pm_graph import (
    answer_pm_questions,
    confirm_prd,
    get_pm_graph_state,
    list_pm_threads,
    run_pm_stream,
)
from src.agent.chaos_graph import get_chaos_state, list_chaos_threads, run_chaos_stream
from src.agent.testpilot_agent import (
    analyze_defect_stream,
    generate_scripts_stream,
    generate_test_cases_stream,
)

load_dotenv()


@asynccontextmanager
async def lifespan(app: FastAPI):  # noqa: ARG001
    yield


app = FastAPI(
    title="PM + Chaos Agents API",
    version="1.0.0",
    description="产品经理智能体 + 混沌测试智能体",
    lifespan=lifespan,
)

app.mount("/ui", StaticFiles(directory="src/ui"), name="ui")


# ═══════════════════════════════════════════════════════════
#  Request / Response Models
# ═══════════════════════════════════════════════════════════

class PMRequest(BaseModel):
    query: str
    thread_id: str | None = None


class PMAnswerRequest(BaseModel):
    answers: list[str]


class PMConfirmRequest(BaseModel):
    approve: bool = True
    feedback: str | None = None


class ChaosRunRequest(BaseModel):
    target_agent: str = "pm"
    external_config: dict | None = None
    test_cases: list[dict] = []
    chaos_scenarios: list[str] = ["text_noise"]
    max_response_time_ms: int = 30000
    tool_whitelist: list[str] = []
    blocked_outputs: list[str] = []
    business_rules: str = ""
    concurrency: int = 3
    timeout_per_case: int = 60


# ═══════════════════════════════════════════════════════════
#  PM Agent Routes
# ═══════════════════════════════════════════════════════════

@app.post("/api/pm/stream")
async def stream_pm_api(request: PMRequest):
    return StreamingResponse(
        run_pm_stream(request.query, request.thread_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/pm/answer/{thread_id}")
async def answer_pm_api(thread_id: str, request: PMAnswerRequest):
    return StreamingResponse(
        answer_pm_questions(thread_id, request.answers),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/pm/confirm/{thread_id}")
async def confirm_pm_api(thread_id: str, request: PMConfirmRequest):
    return StreamingResponse(
        confirm_prd(thread_id, request.approve, request.feedback),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/pm/state/{thread_id}")
async def get_pm_state_api(thread_id: str):
    state = await get_pm_graph_state(thread_id)
    if state is None:
        raise HTTPException(status_code=404, detail="Thread not found")
    return state


@app.get("/api/pm/download/{thread_id}")
async def download_prd(thread_id: str):
    from fastapi.responses import Response
    state = await get_pm_graph_state(thread_id)
    if not state or not state.get("prd_document"):
        raise HTTPException(status_code=404, detail="PRD not found")

    return Response(
        content=state["prd_document"],
        media_type="text/markdown",
        headers={"Content-Disposition": f"attachment; filename=PRD_{thread_id[:8]}.md"},
    )


@app.get("/api/pm/threads")
async def list_pm_threads_api(limit: int = Query(30, ge=1, le=100)):
    """获取 PM Agent 最近对话列表"""
    return await list_pm_threads(limit)


# ═══════════════════════════════════════════════════════════
#  Chaos Testing Routes
# ═══════════════════════════════════════════════════════════

@app.post("/api/chaos/run")
async def chaos_run_api(request: ChaosRunRequest):
    """Run chaos test with SSE streaming progress."""
    config = request.model_dump()
    return StreamingResponse(
        run_chaos_stream(config),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/api/chaos/state/{thread_id}")
async def chaos_state_api(thread_id: str):
    """Get chaos test state for a thread."""
    state = await get_chaos_state(thread_id)
    if state is None:
        raise HTTPException(status_code=404, detail="Thread not found")
    return state


@app.get("/api/chaos/report/{thread_id}")
async def chaos_report_api(thread_id: str):
    """Get chaos test report for a thread."""
    state = await get_chaos_state(thread_id)
    if not state or not state.get("summary"):
        raise HTTPException(status_code=404, detail="Report not found")
    return {
        "thread_id": thread_id,
        "summary": state["summary"],
        "final_result": state.get("final_result"),
    }


@app.get("/api/chaos/threads")
async def chaos_threads_api(limit: int = Query(30, ge=1, le=100)):
    """List recent chaos test threads."""
    return await list_chaos_threads(limit)


# ═══════════════════════════════════════════════════════════
#  AI-TestPilot Routes
# ═══════════════════════════════════════════════════════════

class ExportCasesRequest(BaseModel):
    cases: list[dict]
    format: str = "xlsx"


class GenerateScriptsRequest(BaseModel):
    cases: list[dict]
    script_type: str = "pytest"


class AnalyzeDefectRequest(BaseModel):
    error_log: str
    code_context: str = ""


@app.post("/api/testpilot/generate-cases")
async def testpilot_generate_cases(
    file: UploadFile | None = File(None),
    url: str = Form(""),
    extra_context: str = Form(""),
):
    """Generate test cases from uploaded document or URL."""
    document_text = ""
    if file:
        content = await file.read()
        try:
            document_text = content.decode("utf-8")
        except UnicodeDecodeError:
            document_text = content.decode("gbk", errors="ignore")
    elif url:
        document_text = url

    async def event_stream():
        async for event in generate_test_cases_stream(document_text, extra_context):
            yield f"event: {event['event']}\ndata: {event['data']}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/testpilot/generate-scripts")
async def testpilot_generate_scripts(request: GenerateScriptsRequest):
    """Generate automation scripts from test cases."""
    async def event_stream():
        async for event in generate_scripts_stream(request.cases, request.script_type):
            yield f"event: {event['event']}\ndata: {event['data']}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/testpilot/analyze-defect")
async def testpilot_analyze_defect(request: AnalyzeDefectRequest):
    """Analyze defect from error log."""
    async def event_stream():
        async for event in analyze_defect_stream(request.error_log, request.code_context):
            yield f"event: {event['event']}\ndata: {event['data']}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/testpilot/export-cases")
async def testpilot_export_cases(request: ExportCasesRequest):
    """Export test cases to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header style
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["用例ID", "所属模块", "用例标题", "前置条件", "操作步骤", "预期结果", "优先级"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row, case in enumerate(request.cases, 2):
        steps = case.get("steps", [])
        steps_text = "\n".join(steps) if isinstance(steps, list) else str(steps)
        values = [
            case.get("id", ""),
            case.get("module", ""),
            case.get("title", ""),
            case.get("precondition", ""),
            steps_text,
            case.get("expected", ""),
            case.get("priority", "P2"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # Column widths
    widths = [12, 14, 30, 25, 40, 35, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"},
    )


@app.get("/")
async def root():
    return FileResponse("src/ui/index.html")


@app.get("/api/health")
async def health_check():
    return {"status": "healthy"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "src.api.main:app",
        host=os.getenv("API_HOST", "0.0.0.0"),
        port=int(os.getenv("API_PORT", "8001")),
        reload=True,
    )